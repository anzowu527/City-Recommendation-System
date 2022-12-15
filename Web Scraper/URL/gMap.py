#this file does not scrape website, and is for generating the urls for the city location in google map
from openpyxl import load_workbook

INDEX_URL = 'https://www.google.com/maps/place/{city},+CA'

cities = ['Agoura+Hills','Alhambra','Arcadia','Artesia','Avalon','Azusa','Baldwin+Park','Bell','Bell+Gardens','Bell+flower',
          'Beverly+Hills','Bradbury','Burbank','Calabasas','Carson','Cerritos','City+of+Industry','Claremont','Commerce','Compton','Covina','Cudahy','Culver+City',
          'Diamond+Bar','Downey','Duarte','El+Monte','El+Segundo','Gardena','Glendale','Glendora','Hawaiian+Gardens','Hawthorne','Hermosa+Beach','Hidden+Hills','Huntington+Park','Inglewood',
          'Irwindale','La+Canada+Flintridge','La+Habra+Heights','La+Mirada','La+Puente','La+Verne','Lakewood','Lancaster','Lawndale','Lomita','Long+Beach','Los+Angeles','Lynwood',
          'Malibu','Manhattan+Beach','Maywood','Monrovia','Montebello','Monterey+Park','Norwalk','Palmdale','Palos+Verdes+Estates','Paramount',
          'Pasadena','Pico+Rivera','Pomona','Rancho+Palos+Verdes','Redondo+Beach','Rolling+Hills','Rolling+Hills+Estates','Rosemead','San+Dimas','San+Fernando',
          'San+Gabriel','San+Marino','Santa+Clarita','Santa+Fe+Springs','Santa+Monica','Sierra+Madre','Signal+Hill','South+El+Monte','South+Gate','South+Pasadena','Temple+City',
          'Torrance','Vernon','Walnut','West+Covina','West+Hollywood','Westlake+Village','Whittier']

print(len(cities))

def main():
    d = 1
    url = []
    for i in range(0,len(cities)):
        url.append(INDEX_URL.format(city=cities[i]))

    wb = load_workbook(filename='sURL.xlsx')
    ws = wb.active

    for i in range(len(url)):
        ws.cell(d+1, 2).value = url[i]
        d += 1
    wb.save('sURL.xlsx')

    for i in range(len(url)):
        print(url[i])


main()

