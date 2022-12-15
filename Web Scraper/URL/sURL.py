#this py file is used to get the urls from the school websites
#the sURL.xlsx file under the URL folder is already been created

#the schoolInfo.xlsx is generated from th USNews.py beforehand, and I have it pulled to the URL folder for retrieving the school names

import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

INDEX_URL = 'https://www.google.com/search?q={sName}'

def main():
    #createExcel1()
    wb = load_workbook(filename='sURL.xlsx')
    ws = wb.active
    #remember to change the path if you have a different path :)
    df = pd.read_excel(f'../URL/schoolInfo.xlsx', sheet_name='Sheet', usecols="B")
    sName =[]
    #getting the school name for parsing
    for i in range(len(df)):
        sName.append(df.values[i][0])
    print(sName)
    d = 1

    for i in range(0,len(sName)):
        browser = webdriver.Chrome()
        browser.get(INDEX_URL.format(sName=sName[i]))
        #need to wait more time for the 'website' button to appear
        browser.implicitly_wait(5)
        browser.implicitly_wait(10)
        time.sleep(3)
        url = []
        #make sure the button exist in the page; it exists for most of the time, but there are still a few cases with school that does not have the button, which will require some manual process
        size = browser.find_elements(By.CLASS_NAME, 'ab_button')

        if len(size) == 0:
            print('hi')
            url.append('N/A')
            browser.close()

        else:
            browser.implicitly_wait(5)
            browser.implicitly_wait(10)
            browser.find_element(By.CLASS_NAME, 'ab_button').click()
            url.append(browser.current_url)
            browser.close()

        for i in range(len(url)):
            ws.cell(d + 1, 1).value = url[i]
            d += 1

        wb.save('sURL.xlsx')


if __name__ == '__main__':
    main()



