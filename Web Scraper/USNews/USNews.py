#this py file is responsible to scrape down the school related information form the USNews.com
#I scraped more information as I need in case for future usage
#when the file starts running, it will automatically generate a schoolInfo.xlsx file under the USNews folder
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from pyquery import PyQuery as pq


INDEX_URL = 'https://www.usnews.com/education/best-high-schools/california/districts/arcadia-unified-school-district/arcadia-high-school-{number}'

cities = ['Agoura Hills','Alhambra','Arcadia','Artesia','Avalon','Azusa','Baldwin Park','Bell','Bell Gardens','Bellflower',
          'Beverly Hills','Bradbury','Burbank','Calabasas','Carson','Cerritos','City of Industry','Claremont','Commerce','Compton','Covina','Cudahy','Culver City',
          'Diamond Bar','Downey','Duarte','El Monte','El Segundo','Gardena','Glendale','Glendora','Hawaiian Gardens','Hawthorne','Hermosa Beach','Hidden Hills','Huntington Park','Inglewood','Irwindale',
          'La Cañada Flintridge','La Habra Heights','La Mirada','La Puente','La Verne','Lakewood','Lancaster','Lawndale','Lomita','Long Beach','Los Angeles','Lynwood',
          'Malibu','Manhattan Beach','Maywood','Monrovia','Montebello','Monterey Park','Norwalk','Palmdale','Palos Verdes Estates','Paramount',
          'Pasadena','Pico Rivera','Pomona','Rancho Palos Verdes','Redondo Beach','Rolling Hills','Rolling Hills Estates','Rosemead','San Dimas','San Fernando',
          'San Gabriel','San Marino','Santa Clarita','Santa Fe Springs','Santa Monica','Sierra Madre','Signal Hill','South El Monte','South Gate','South Pasadena','Temple City',
          'Torrance','Vernon','Walnut','West Covina','West Hollywood','Westlake Village','Whittier']

dict = [2434,1663,1787,1705,2498,1801,1805,2516,2801,1823,1833,2073,1862,2435,2519,1706,2308,1962,201004,1994,2032,140705,2035,3634,2071,
        2073,2117,2124,2540,2277,2281,2505,1916,1542,2530,2550,2352,1802,2421,2243,2899,2304,1847,2492,1769,1919,2571,2494,2562,2725,3400,1542,2645,
        2798,2802,2807,2900,1767,2999,3005,3015,2122,3061,2999,1565,2998,2998,2119,1849,2585,1665,3334,3670,3659,3402,3015,153241,2115,
        2587,3484,3555,3563,2664,3633,3645,2534,2003,3660]

def createExcel():
    wb = openpyxl.Workbook()
    wb.save(f'../USNews/schoolInfo.xlsx')

    wb = load_workbook(filename='schoolInfo.xlsx')
    ws = wb.active

    ws.cell(1, 1).value = 'city'
    ws.cell(1, 2).value = 'sName'
    ws.cell(1, 3).value = 'californiaRankings'
    ws.cell(1, 4).value = 'tEnroll'
    ws.cell(1, 5).value = 'Ratio'
    ws.cell(1, 6).value = 'white'
    ws.cell(1, 7).value = 'asian'
    ws.cell(1, 8).value = 'hispanic'
    ws.cell(1, 9).value = 'black'
    ws.cell(1, 10).value = 'American Indian/Alaska Native'
    ws.cell(1, 11).value = 'Native Hawaiian/Pacific Islander'
    ws.cell(1, 12).value = 'male'
    ws.cell(1, 13).value = 'female'
    ws.cell(1, 14).value = 'District#School'
    ws.cell(1, 15).value = 'District#student'
    ws.cell(1, 16).value = 'address'

    wb.save('schoolInfo.xlsx')

def main():
    createExcel()
    #loop through every urls one by one
    d = 1

    for i in range(0,len(dict)):
        datas = []
        data = []

        browser = webdriver.Chrome()
        browser.get(INDEX_URL.format(number=dict[i]))
        wb = load_workbook(filename='schoolInfo.xlsx')
        ws = wb.active

        city = cities[i]
        data.append(city)
        #parsing url
        doc = pq(browser.page_source)
        name = doc('.knjdTp')
        sName = name('.hcuouP').text()
        data.append(sName)
        enroll = browser.find_element(By.CSS_SELECTOR, 'strong[data-test-id="ccd_member"]').text
        ratio = browser.find_element(By.CSS_SELECTOR, 'strong[data-test-id="student_teacher_ratio_rounded"]').text

        #scraping the race and the corresponding percentage
        race = doc('.cuHcPI')
        perc = []
        races = []
        for k in race.items():
            g = 0
            percent = k('.fPxKtP').text()
            race = k('.hWVZdP').text()
            perc.append(percent)
            races.append(race)
            g += 1
        #print('percent = ', perc)
        #print('race = ', races)

        #getting rid of the %
        for i in range(len(perc)):
            a = perc[i].split('%')
            perc[i] = a[0]
        print(perc)
        white=''
        asian=''
        hispanic=''
        black=''
        am_native=''
        hawaii=''
        for i in range(len(races)):
            if races[i] == 'White':
                white = perc[i]
            if races[i] == 'Asian':
                asian= perc[i]
            if races[i] == 'Hispanic':
                hispanic= perc[i]
            if races[i] == 'Black':
                black= perc[i]
            if races[i] == 'American Indian/Alaska Native':
                am_native= perc[i]
            if races[i] == 'Native Hawaiian/Pacific Islander':
                hawaii= perc[i]

        #not all the races will be listed out in USNews
        #if the school does not cantain the specific race, we set it to 0%
        if white == '':
            white= '0'
        if asian == '':
            asian= '0'
        if hispanic == '':
            hispanic= '0'
        if black == '':
            black= '0'
        if am_native == '':
            am_native= '0'
        if hawaii == '':
            hawaii= '0'

        male = browser.find_element(By.XPATH, '//*[@id="students_teachers_section"]/react-trigger[2]/div/div[2]/div/div/div/div/div/div[2]/div/div[2]/b').text
        female = browser.find_element(By.XPATH, '//*[@id="students_teachers_section"]/react-trigger[2]/div/div[2]/div/div/div/div/div/div[2]/div/div[1]/b').text

        districtS = browser.find_element(By.XPATH, '//*[@id="district_section"]/div/div[1]/p[2]/strong').text
        districtSchool = int(districtS)

        if districtSchool > 1:
            districtTotal = browser.find_element(By.XPATH, '//*[@id="district_section"]/div/div[2]/p[2]/strong').text
        else:
            districtTotal = enroll

        address = browser.find_element(By.CLASS_NAME, 'Paragraph-sc-1iyax29-0.dDjpVm').text
        address = address[0:len(address)-16]

        unrank = browser.find_element(By.CLASS_NAME, 'gwrUrw').text
        unrank = unrank.split()

        #some time the school is unranked
        if unrank[len(unrank) - 1] == 'Unranked':
            rank = 'unranked'
            data.append(rank)

        else:
            # parse the rank string and get the California ranking & take away the # symbol
            # it is possible for the string 'iRlwVS' to change over time - updated on 14/12/2022
            # this string can be found when you open the inspect page and locate the section 'All Rankings', and copy the string after the last space in the <>
            rank = browser.find_element(By.CLASS_NAME, 'iRlwVS').text
            rank = rank.split()
            rank = rank[4].split('#')
            data.append(rank[1])

        data.append(enroll)
        data.append(ratio)
        data.append(white)
        data.append(asian)
        data.append(hispanic)
        data.append(black)
        data.append(am_native)
        data.append(hawaii)
        data.append(male)
        data.append(female)
        data.append(districtSchool)
        data.append(districtTotal)
        data.append(address)
        datas.append(data)
        #print(data)

        #write the information from the website to the excel
        for data in datas:
            for i in range(1,17):
                ws.cell(d + 1, i).value = data[i-1]

        d += 1
        wb.save('schoolInfo.xlsx')
        browser.quit()

main()

