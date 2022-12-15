#this py file scrapes the house price information
#I also scraped the information other then the house price in case I need them in the future
# this file will auto-generate the saleHome1.xlsx to store the information

import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from pyquery import PyQuery as pq
browser = webdriver.Chrome()

INDEX_URL = 'https://www.redfin.com/city/{number}/CA/Arcadia'

def createExcel1():
    wb = openpyxl.Workbook()
    wb.save(f'../redfin/saleHome1.xlsx')
    wb = load_workbook('saleHome1.xlsx')
    ws = wb.active

    ws.cell(1, 1).value = 'city'
    ws.cell(1, 2).value = 'price'
    ws.cell(1, 3).value = 'bed'
    ws.cell(1, 4).value = 'baths'
    ws.cell(1, 5).value = 'sqFt'
    ws.cell(1, 6).value = 'address'

    wb.save('saleHome1.xlsx')

def scrape_url(url):
    browser.get(url)
    browser.implicitly_wait(5)

def parse_url(html):
    doc = pq(html)
    items = doc('.defaultSplitMapListView')
    datas = []

    for i in items.items():
        data = []
        city = browser.find_element(By.XPATH, '//*[@id="headerUnifiedSearch"]/div/form/div/div/div[1]/div/div[1]').text
        price = i('.homecardV2Price').text()
        #using copy selector
        bed = i('.HomeStatsV2>div:nth-child(1)').text()
        baths = i('.HomeStatsV2>div:nth-child(2)').text()
        sqFt = i('.HomeStatsV2>div:nth-child(3)').text()
        address = i('.primaryLine').text()

        data.append(city)
        data.append(price)
        data.append(bed)
        data.append(baths)
        data.append(sqFt)
        data.append(address)

        datas.append(data)

    return datas

dict1= [77, 208, 645,761,878,921,998,1285,1331,1328,1669,2048,2320,2471,2990,3268,9288,3578,3870,3890,4367,4539,4557,4957,5092,5152,5726,5771,7193,7646,7650,
        8292,8306,8536,8583,9186,9307,9373,9895,9982,10208,10290,10417,10167,10233,10433,10810,10940,11203,11368,11532,11576,11864,12449,12502,12524,13533,14292,14343,14399,14498,14737,
        15047,15404,15507,16076,16085,16137,16923,16938,17164,17481,17676,17691,17882,18335,18347,18603,18624,18661,19707,20094,20486,20628,20753,20771,20777,20869]

dict2=[208,77]
'''dict1= [77]

dict3=[645]'''
def main():
    time.sleep(10)
    createExcel1()
    d = 1
    cookie = 0
    for i in range(0, len(dict1)):
        browser.get(INDEX_URL.format(number=dict1[i]))
        browser.implicitly_wait(5)
        # getting the number of pages by parsing the page text
        # since there are no patterns in the html for the page buttons
        # parsing it and get the last page number: ex[Viewing page 3 of 9]
        pages = browser.find_element(By.CLASS_NAME, 'pageText').text
        pages = pages.split()
        page = int(pages[4])

        wb = load_workbook(filename='saleHome1.xlsx')
        ws = wb.active
        for p in range(0, page):
            browser.implicitly_wait(20)
            datas = parse_url(browser.page_source)
            # adding the data from the website to the excel
            for data in datas:
                ws.cell(d+1, 1).value = data[0]
                ws.cell(d+1, 2).value = data[1]
                ws.cell(d+1, 3).value = data[2]
                ws.cell(d+1, 4).value = data[3]
                ws.cell(d+1, 5).value = data[4]
                ws.cell(d+1, 6).value = data[5]
                d += 1

            wb.save('saleHome1.xlsx')

            # when it is the first time running the program and open the webpage, it will have a pop out asking to accept cookies.
            # We have to click the 'accept cookie' once in order to progress
            # the XPATH might change when the developer modified the website, and the new XPATH can be found in the inspect page by locating the element that we want to click on
            # XPATH Last update: 14/12/2022
            if cookie == 0:
                browser.find_element(By.XPATH, '//*[@id="onetrust-accept-btn-handler"]').click()
                cookie = 1
            #if there are multiple page, click the button to the next page until it is the end of the pages
            if p != page-1:
                browser.find_element(By.XPATH, '//*[@id="results-display"]/div[4]/div/div/div[3]/button[2]').click()
                browser.implicitly_wait(20)

    #sometime when I run all the cities at once, error will pop up toward the end (usually no errors), and it took a long time to scrape from the beginning
    #If this situation happens, I will split the cities in many groups, and scrape them seperately

if __name__ == '__main__':
    main()
    browser.close()

