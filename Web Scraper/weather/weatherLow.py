# this file is for scraping the city average lowest temperatures throughout a year
# this file will auto-generate the weatherLow.xlsx to store the information
# It is possible to write the scraping code for scraping the highest and lowest temperature in the same file, but I seperate it for saving time by running these two file parallel
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

INDEX_URL = 'https://weatherspark.com/y/{index}/Average-Weather-in-Arcadia-California-United-States-Year-Round'

def createExcel1():
    wb = openpyxl.Workbook()
    wb.save(f'../weather/weatherLow.xlsx')
    wb = load_workbook('weatherLow.xlsx')
    ws = wb.active

    ws.cell(1, 1).value = 'city'
    ws.cell(1, 2).value = 'Jan'
    ws.cell(1, 3).value = 'Feb'
    ws.cell(1, 4).value = 'March'
    ws.cell(1, 5).value = 'Apr'
    ws.cell(1, 6).value = 'may'
    ws.cell(1, 7).value = 'June'
    ws.cell(1, 8).value = 'July'
    ws.cell(1, 9).value = 'Aug'
    ws.cell(1, 10).value = 'Sep'
    ws.cell(1, 11).value = 'Oct'
    ws.cell(1, 12).value = 'Nov'
    ws.cell(1, 13).value = 'Dec'

    wb.save('weatherLow.xlsx')

index = [1676,1678,1680,1613,1614,1922,1923,1615,1616,1617,1682,1928,1684,1685,1618,1619,1850,1933,1690,1620,1935,1621,1691,
         1937,1624,1938,1695,1626,1627,1697,1941,1628,1629,1630,1699,1631,1632,1946,1700,1857,1633,1947,1948,1636,1701,1637,1639,1640,1705,1642,
         1706,1643,1645,1955,1709,1710,1646,1717,1647,1648,1718,1649,1963,1650,1651,1653,1653,1722,1970,1723,1724,1725,1726,1656,1727,1730,1658,1732,1659,1733,1735,1661,1662,1976,1977,1743,1744,1668]
print(len(index))

def main():
    createExcel1()
    wb = load_workbook(filename='weatherLow.xlsx')
    ws = wb.active
    d = 1
    for i in range(0, len(index)):
        high = []
        browser = webdriver.Chrome()
        browser.get(INDEX_URL.format(index=index[i]))
        h = browser.find_elements(By.XPATH, '//*[@id="Report-Content"]/div[2]/div[2]/div[1]/table/tbody/tr/td/table/tbody/tr[3]')
        for value in h:
            high.append(value.text)

        high = high[0].split(' ')

        #extract only the number
        high = high[1:len(high)]
        for i in range(len(high)):
            a = high[i].split('°F')
            high[i] = a[0]

        #get the corresponding city name
        cName = []
        for value in browser.find_elements(By.XPATH, '//*[@id="Sidebar-Container"]/h4'):
            print(value.text)
            cName.append(value.text)

        for i in range(len(high)):
            cName.append(high[i])

        print(cName)

        datas = []
        datas.append(cName)

        for data in datas:
            for i in range(1,14):
                ws.cell(d + 1, i).value = data[i-1]
            d += 1

        wb.save('weatherLow.xlsx')
        browser.quit()


if __name__ == '__main__':
    main()
